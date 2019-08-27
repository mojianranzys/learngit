# coding=utf_8_sig
import pandas as pd
from openpyxl import load_workbook
import os
import argparse
import openpyxl

parser = argparse.ArgumentParser()
parser.add_argument("-d", "--dir", nargs='?', action="store" , type=str, help="annokb sample directory")
args = parser.parse_args()

path = args.dir
os_chdir = path + 'Combine_mutscan_fusion_virus/'
def _excelAddSheet(csv_name, excelWriter, sheet_name):
    dataframe = pd.read_csv(csv_name)
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=None)
    excelWriter.close()
pd.DataFrame().to_excel(os.path.join(os_chdir, 'mutscan_fusion_virus.xlsx'))
excelWriter = pd.ExcelWriter(os.path.join(os_chdir, 'mutscan_fusion_virus.xlsx'),engine='openpyxl')

sample = path.split("/")[-2]
print(sample)
mutscan_csv = os_chdir + sample + '_mutscan_result.csv'
fusion_csv = os_chdir + sample + '_fusion_result.csv'
virus_csv = os_chdir + sample + '_virus_result.csv'
name = ["mutscan","fusion","virus"]
list1 = [mutscan_csv,fusion_csv,virus_csv]
for i in range(len(list1)):
    sheet_name = name[i][:10]
    print(sheet_name)
    _excelAddSheet(list1[i], excelWriter,sheet_name)


sExcelFile=os_chdir + 'mutscan_fusion_virus.xlsx'
wb = openpyxl.load_workbook(sExcelFile)
ws = wb["Sheet1"]
wb.remove(ws)
wb.save(sExcelFile)
