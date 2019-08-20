# coding=utf_8_sig
import pandas as pd
from openpyxl import load_workbook
import os
import openpyxl
os.chdir('C:/Users/user/Desktop/')
output_dir = 'C:/Users/user/Desktop/'
def _excelAddSheet(csv_name, excelWriter, sheet_name):
    dataframe = pd.read_csv(csv_name)
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=None)
    excelWriter.close()
pd.DataFrame().to_excel(os.path.join(output_dir, 'mutscan_fusion_virus.xlsx'))
excelWriter = pd.ExcelWriter(os.path.join(output_dir, 'mutscan_fusion_virus.xlsx'),engine='openpyxl')
# excel必需已经存在，因此先建立一个空的sheet
list1 = ['S018_SZ20190726043PES-6_ffpedna_451plus_30863_mutscan_result.csv','S055_SZ20190725009WHB-0_cfdna_451plus_30756_fusion_result.csv','virus_tumor_normal_result.csv']
name = ["mutscan","fusion","virus"]
for i in range(len(list1)):
    sheet_name = name[i][:10]
    _excelAddSheet(list1[i], excelWriter,sheet_name)

sExcelFile="C:/Users/user/Desktop/mutscan_fusion_virus.xlsx"
wb = openpyxl.load_workbook(sExcelFile)
ws = wb["Sheet1"]
wb.remove(ws)
wb.save(sExcelFile)
